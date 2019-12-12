import openpyxl
import random

wb = openpyxl.load_workbook('D:\qfc_asset.xlsx')

sheets = wb.active

max_row = sheets.max_row
# max_col = sheets.max_column

# print(type(max_row))

for erow in range(1,max_row):
    cvalue = random.random()
    sheets.cell(row = erow, column = 24).value = str(float('%.4f' % cvalue))

wb.save('D:\qfc_asset.xlsx')

print("Done!!!")